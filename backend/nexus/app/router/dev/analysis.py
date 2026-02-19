from typing import List, Dict, Any
from datetime import datetime, timedelta
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from app.database.database import projects_coll
from app.database.schemas.project import Project
from app.database.schemas.department_data import EPCData
from app.database.schemas.security import AccessTokenData
from app.security.oauth2 import get_current_user_data

"""
Endpoint: /dev/analysis
Purpose:
  - Analysis and reporting endpoints for project data
  - Categories: drafting, engineering, plat, county (permitting)
  - Periods: yearly (/2025, /2026), monthly (/2026/1 … /2026/12), weekly (/current-week)
"""

router = APIRouter(prefix="/dev/analysis")

ALLOWED_ROLES = [299, 999]

MONTH_NAMES = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December",
}


# ---------------------------------------------------------------------------
# Auth dependency
# ---------------------------------------------------------------------------

def require_analysis_roles(user_data: AccessTokenData = Depends(get_current_user_data)):
    user_roles = user_data.roles
    if not any(role in user_roles for role in ALLOWED_ROLES):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="You don't have access to this resource. Required roles: 299 or 999",
        )
    return user_data


# ---------------------------------------------------------------------------
# Data helpers
# ---------------------------------------------------------------------------

def _days_between(start: datetime | None, end: datetime | None) -> int | None:
    if start and end:
        return max(0, round((end - start).total_seconds() / 86400))
    return None


def _fmt_date(dt: datetime | None) -> str | None:
    return dt.strftime("%Y-%m-%d") if dt else None


def _fetch_all_project_infos() -> List[Dict[str, Any]]:
    """
    Fetch and parse all projects from MongoDB into flat dicts.
    Only projects with a contract_date are included.
    """
    result: List[Dict[str, Any]] = []

    for doc in projects_coll.find():
        project_raw = {k: v for k, v in doc.items() if k != "_id"}
        try:
            project: Project = Project(**project_raw)
            epc: EPCData = EPCData(**project_raw["teclab_data"]["epc_data"])

            if not epc.contract_date:
                continue

            result.append({
                # identity
                "project_id":    project.project_info.project_id,
                "community":     project.project_info.community,
                "section":       project.project_info.section,
                "lot":           project.project_info.lot_number,
                # contract
                "contract_date": epc.contract_date,
                "contract_type": epc.contract_type,
                "product_name":  epc.product_name,
                "elevation_name": epc.elevation_name,
                # drafting
                "drafting_drafter":     epc.drafting_drafter,
                "drafting_assigned_on": epc.drafting_assigned_on,
                "drafting_finished":    epc.drafting_finished,
                "drafting_days":        _days_between(epc.drafting_assigned_on, epc.drafting_finished),
                "drafting_notes":       epc.drafting_notes,
                # engineering
                "engineering_engineer": epc.engineering_engineer,
                "engineering_sent":     epc.engineering_sent,
                "engineering_received": epc.engineering_received,
                "engineering_days":     _days_between(epc.engineering_sent, epc.engineering_received),
                "engineering_notes":    epc.engineering_notes,
                # plat
                "plat_engineer":  epc.plat_engineer,
                "plat_sent":      epc.plat_sent,
                "plat_received":  epc.plat_received,
                "plat_days":      _days_between(epc.plat_sent, epc.plat_received),
                "plat_notes":     epc.plat_notes,
                # permitting / county
                "permitting_county":    epc.permitting_county_name,
                "permitting_submitted": epc.permitting_submitted,
                "permitting_received":  epc.permitting_received,
                "permitting_days":      _days_between(epc.permitting_submitted, epc.permitting_received),
                "permitting_notes":     epc.permitting_notes,
            })
        except Exception as e:
            print(f"ERROR parsing project: {e}")
            continue

    return result


def _stats(days_list: List[int]) -> Dict[str, Any]:
    if not days_list:
        return {"avg": None, "min": None, "max": None}
    return {
        "avg": round(sum(days_list) / len(days_list), 1),
        "min": min(days_list),
        "max": max(days_list),
    }


def _build_report(projects: List[Dict[str, Any]], period_label: Dict[str, Any]) -> Dict[str, Any]:
    """
    Build the full report dict from a filtered, sorted project list.
    """
    # --- summary ---
    def _count(key): return sum(1 for p in projects if p[key])

    drafting_days   = [p["drafting_days"]   for p in projects if p["drafting_days"]   is not None]
    engineering_days = [p["engineering_days"] for p in projects if p["engineering_days"] is not None]
    plat_days       = [p["plat_days"]       for p in projects if p["plat_days"]       is not None]
    permitting_days = [p["permitting_days"] for p in projects if p["permitting_days"] is not None]

    ds = _stats(drafting_days)
    es = _stats(engineering_days)
    ps = _stats(plat_days)
    cs = _stats(permitting_days)

    summary = {
        "total_projects": len(projects),
        # drafting
        "drafting_assigned":        _count("drafting_assigned_on"),
        "drafting_complete":        _count("drafting_finished"),
        "drafting_avg_days":        ds["avg"],
        "drafting_min_days":        ds["min"],
        "drafting_max_days":        ds["max"],
        # engineering
        "engineering_sent":         _count("engineering_sent"),
        "engineering_received":     _count("engineering_received"),
        "engineering_avg_days":     es["avg"],
        "engineering_min_days":     es["min"],
        "engineering_max_days":     es["max"],
        # plat
        "plat_sent":                _count("plat_sent"),
        "plat_received":            _count("plat_received"),
        "plat_avg_days":            ps["avg"],
        "plat_min_days":            ps["min"],
        "plat_max_days":            ps["max"],
        # permitting / county
        "permitting_submitted":     _count("permitting_submitted"),
        "permitting_received":      _count("permitting_received"),
        "permitting_avg_days":      cs["avg"],
        "permitting_min_days":      cs["min"],
        "permitting_max_days":      cs["max"],
    }

    # --- by drafter ---
    by_drafter: Dict[str, Any] = {}
    for p in projects:
        name = p["drafting_drafter"] or "Unassigned"
        bucket = by_drafter.setdefault(name, {"count": 0, "days": []})
        bucket["count"] += 1
        if p["drafting_days"] is not None:
            bucket["days"].append(p["drafting_days"])
    drafter_breakdown = {
        name: {"total_projects": b["count"], "complete": len(b["days"]), **_stats(b["days"])}
        for name, b in by_drafter.items()
    }

    # --- by engineer ---
    by_engineer: Dict[str, Any] = {}
    for p in projects:
        name = p["engineering_engineer"] or "Unassigned"
        bucket = by_engineer.setdefault(name, {"count": 0, "days": []})
        bucket["count"] += 1
        if p["engineering_days"] is not None:
            bucket["days"].append(p["engineering_days"])
    engineer_breakdown = {
        name: {"total_projects": b["count"], "complete": len(b["days"]), **_stats(b["days"])}
        for name, b in by_engineer.items()
    }

    # --- by plat engineer ---
    by_plat: Dict[str, Any] = {}
    for p in projects:
        name = p["plat_engineer"] or "Unassigned"
        bucket = by_plat.setdefault(name, {"count": 0, "days": []})
        bucket["count"] += 1
        if p["plat_days"] is not None:
            bucket["days"].append(p["plat_days"])
    plat_breakdown = {
        name: {"total_projects": b["count"], "complete": len(b["days"]), **_stats(b["days"])}
        for name, b in by_plat.items()
    }

    # --- by county ---
    by_county: Dict[str, Any] = {}
    for p in projects:
        name = p["permitting_county"] or "Unknown"
        bucket = by_county.setdefault(name, {"count": 0, "days": []})
        bucket["count"] += 1
        if p["permitting_days"] is not None:
            bucket["days"].append(p["permitting_days"])
    county_breakdown = {
        name: {"total_projects": b["count"], "complete": len(b["days"]), **_stats(b["days"])}
        for name, b in by_county.items()
    }

    return {
        **period_label,
        "summary":      summary,
        "by_drafter":   drafter_breakdown,
        "by_engineer":  engineer_breakdown,
        "by_plat":      plat_breakdown,
        "by_county":    county_breakdown,
        "projects":     projects,
    }


# ---------------------------------------------------------------------------
# Period helpers
# ---------------------------------------------------------------------------

# Start dates used to determine if a lot "belongs" to a period.
# A lot is included if ANY of these dates fall within the period.
_WORK_START_FIELDS = [
    "drafting_assigned_on",
    "engineering_sent",
    "plat_sent",
    "permitting_submitted",
]


def _any_in_period(p: Dict[str, Any], start: datetime, end: datetime) -> bool:
    """Return True if any work-start date falls within [start, end)."""
    for field in _WORK_START_FIELDS:
        dt = p.get(field)
        if dt is not None:
            dt = dt.replace(tzinfo=None)
            if start <= dt < end:
                return True
    return False


def _report_for_year(year: int) -> Dict[str, Any]:
    period_start = datetime(year, 1, 1)
    period_end   = datetime(year + 1, 1, 1)
    projects = [
        p for p in _fetch_all_project_infos()
        if _any_in_period(p, period_start, period_end)
    ]
    projects.sort(key=lambda p: min(
        (p[f].replace(tzinfo=None) for f in _WORK_START_FIELDS if p[f]),
        default=datetime.max,
    ))
    return _build_report(projects, {"year": year})


def _report_for_month(year: int, month: int) -> Dict[str, Any]:
    period_start = datetime(year, month, 1)
    # advance to first day of next month
    if month == 12:
        period_end = datetime(year + 1, 1, 1)
    else:
        period_end = datetime(year, month + 1, 1)
    projects = [
        p for p in _fetch_all_project_infos()
        if _any_in_period(p, period_start, period_end)
    ]
    projects.sort(key=lambda p: min(
        (p[f].replace(tzinfo=None) for f in _WORK_START_FIELDS if p[f]),
        default=datetime.max,
    ))
    return _build_report(projects, {"year": year, "month": month})


def _report_for_current_week() -> Dict[str, Any]:
    today = datetime.utcnow().date()
    period_start = datetime.combine(today - timedelta(days=today.weekday()), datetime.min.time())
    period_end   = period_start + timedelta(days=7)

    projects = [
        p for p in _fetch_all_project_infos()
        if _any_in_period(p, period_start, period_end)
    ]
    projects.sort(key=lambda p: min(
        (p[f].replace(tzinfo=None) for f in _WORK_START_FIELDS if p[f]),
        default=datetime.max,
    ))
    return _build_report(projects, {
        "week_start": period_start.strftime("%Y-%m-%d"),
        "week_end":   (period_end - timedelta(days=1)).strftime("%Y-%m-%d"),
    })


# ---------------------------------------------------------------------------
# Excel generator
# ---------------------------------------------------------------------------

def _period_title(data: Dict[str, Any]) -> str:
    if "week_start" in data:
        return f"Week of {data['week_start']} – {data['week_end']}"
    if "month" in data:
        return f"{data['year']} {MONTH_NAMES.get(data['month'], data['month'])}"
    return str(data["year"])


def _period_filename(data: Dict[str, Any]) -> str:
    if "week_start" in data:
        return f"week_{data['week_start']}"
    if "month" in data:
        return f"{data['year']}_{MONTH_NAMES.get(data['month'], data['month'])}"
    return str(data["year"])


def generate_excel(data: Dict[str, Any]) -> BytesIO:
    title = _period_title(data)
    wb = Workbook()

    # Shared styles
    header_font      = Font(bold=True, color="FFFFFF")
    header_fill      = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border      = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def style_headers(ws, num_cols: int):
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font      = header_font
            cell.fill      = header_fill
            cell.alignment = header_alignment
            cell.border    = thin_border

    def auto_width(ws):
        for col_idx, col_cells in enumerate(ws.columns, 1):
            width = max((len(str(c.value)) for c in col_cells if c.value), default=0)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(width + 2, 50)

    def write_breakdown_sheet(ws, col1_label: str, breakdown: Dict[str, Any], stat_keys: List[str], stat_headers: List[str]):
        headers = [col1_label, "Total Projects", "Complete"] + stat_headers
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h)
        style_headers(ws, len(headers))
        for r, (name, stats) in enumerate(breakdown.items(), start=2):
            ws.cell(row=r, column=1, value=name)
            ws.cell(row=r, column=2, value=stats["total_projects"])
            ws.cell(row=r, column=3, value=stats["complete"])
            ws.cell(row=r, column=4, value=stats["avg"])
            ws.cell(row=r, column=5, value=stats["min"])
            ws.cell(row=r, column=6, value=stats["max"])
        auto_width(ws)

    summary = data["summary"]

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary["A1"] = f"{title} Analysis Summary"
    ws_summary["A1"].font = Font(bold=True, size=14)
    ws_summary.merge_cells("A1:B1")

    sections = [
        ("DRAFTING", [
            ("Assigned",        summary["drafting_assigned"]),
            ("Complete",        summary["drafting_complete"]),
            ("Avg Days",        summary["drafting_avg_days"]),
            ("Min Days",        summary["drafting_min_days"]),
            ("Max Days",        summary["drafting_max_days"]),
        ]),
        ("ENGINEERING", [
            ("Sent",            summary["engineering_sent"]),
            ("Received",        summary["engineering_received"]),
            ("Avg Days",        summary["engineering_avg_days"]),
            ("Min Days",        summary["engineering_min_days"]),
            ("Max Days",        summary["engineering_max_days"]),
        ]),
        ("PLAT", [
            ("Sent",            summary["plat_sent"]),
            ("Received",        summary["plat_received"]),
            ("Avg Days",        summary["plat_avg_days"]),
            ("Min Days",        summary["plat_min_days"]),
            ("Max Days",        summary["plat_max_days"]),
        ]),
        ("COUNTY / PERMITTING", [
            ("Submitted",       summary["permitting_submitted"]),
            ("Received",        summary["permitting_received"]),
            ("Avg Days",        summary["permitting_avg_days"]),
            ("Min Days",        summary["permitting_min_days"]),
            ("Max Days",        summary["permitting_max_days"]),
        ]),
    ]

    current_row = 3
    for section_title, rows in sections:
        ws_summary.cell(row=current_row, column=1, value=section_title).font = Font(bold=True, size=12)
        current_row += 1
        for label, value in rows:
            ws_summary.cell(row=current_row, column=1, value=label).font = Font(bold=True)
            ws_summary.cell(row=current_row, column=2, value=value)
            current_row += 1
        current_row += 1  # blank row between sections

    ws_summary.cell(row=current_row, column=1, value="Total Projects").font = Font(bold=True)
    ws_summary.cell(row=current_row, column=2, value=summary["total_projects"])
    auto_width(ws_summary)

    # ── Sheet 2: By Drafter ───────────────────────────────────────────────────
    write_breakdown_sheet(
        wb.create_sheet("By Drafter"), "Drafter",
        data["by_drafter"], [], ["Avg Days", "Min Days", "Max Days"],
    )

    # ── Sheet 3: By Engineer ──────────────────────────────────────────────────
    write_breakdown_sheet(
        wb.create_sheet("By Engineer"), "Engineer",
        data["by_engineer"], [], ["Avg Days", "Min Days", "Max Days"],
    )

    # ── Sheet 4: By Plat Engineer ─────────────────────────────────────────────
    write_breakdown_sheet(
        wb.create_sheet("By Plat Engineer"), "Plat Engineer",
        data["by_plat"], [], ["Avg Days", "Min Days", "Max Days"],
    )

    # ── Sheet 5: By County ────────────────────────────────────────────────────
    write_breakdown_sheet(
        wb.create_sheet("By County"), "County",
        data["by_county"], [], ["Avg Days", "Min Days", "Max Days"],
    )

    # ── Sheet 6: All Projects ─────────────────────────────────────────────────
    ws_projects = wb.create_sheet("All Projects")
    project_headers = [
        "Project ID", "Community", "Section", "Lot",
        "Contract Date", "Contract Type", "Product", "Elevation",
        # drafting
        "Drafter", "Drafting Assigned", "Drafting Finished", "Drafting Days", "Drafting Notes",
        # engineering
        "Engineer", "Eng Sent", "Eng Received", "Eng Days", "Eng Notes",
        # plat
        "Plat Engineer", "Plat Sent", "Plat Received", "Plat Days", "Plat Notes",
        # permitting
        "County", "Permit Submitted", "Permit Received", "Permit Days", "Permit Notes",
    ]
    for col, h in enumerate(project_headers, 1):
        ws_projects.cell(row=1, column=col, value=h)
    style_headers(ws_projects, len(project_headers))

    for r, p in enumerate(data["projects"], start=2):
        ws_projects.cell(row=r, column=1,  value=p["project_id"])
        ws_projects.cell(row=r, column=2,  value=p["community"])
        ws_projects.cell(row=r, column=3,  value=p["section"])
        ws_projects.cell(row=r, column=4,  value=p["lot"])
        ws_projects.cell(row=r, column=5,  value=_fmt_date(p["contract_date"]))
        ws_projects.cell(row=r, column=6,  value=p["contract_type"])
        ws_projects.cell(row=r, column=7,  value=p["product_name"])
        ws_projects.cell(row=r, column=8,  value=p["elevation_name"])
        # drafting
        ws_projects.cell(row=r, column=9,  value=p["drafting_drafter"])
        ws_projects.cell(row=r, column=10, value=_fmt_date(p["drafting_assigned_on"]))
        ws_projects.cell(row=r, column=11, value=_fmt_date(p["drafting_finished"]))
        ws_projects.cell(row=r, column=12, value=p["drafting_days"])
        ws_projects.cell(row=r, column=13, value=p["drafting_notes"])
        # engineering
        ws_projects.cell(row=r, column=14, value=p["engineering_engineer"])
        ws_projects.cell(row=r, column=15, value=_fmt_date(p["engineering_sent"]))
        ws_projects.cell(row=r, column=16, value=_fmt_date(p["engineering_received"]))
        ws_projects.cell(row=r, column=17, value=p["engineering_days"])
        ws_projects.cell(row=r, column=18, value=p["engineering_notes"])
        # plat
        ws_projects.cell(row=r, column=19, value=p["plat_engineer"])
        ws_projects.cell(row=r, column=20, value=_fmt_date(p["plat_sent"]))
        ws_projects.cell(row=r, column=21, value=_fmt_date(p["plat_received"]))
        ws_projects.cell(row=r, column=22, value=p["plat_days"])
        ws_projects.cell(row=r, column=23, value=p["plat_notes"])
        # permitting
        ws_projects.cell(row=r, column=24, value=p["permitting_county"])
        ws_projects.cell(row=r, column=25, value=_fmt_date(p["permitting_submitted"]))
        ws_projects.cell(row=r, column=26, value=_fmt_date(p["permitting_received"]))
        ws_projects.cell(row=r, column=27, value=p["permitting_days"])
        ws_projects.cell(row=r, column=28, value=p["permitting_notes"])

    auto_width(ws_projects)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def _excel_response(data: Dict[str, Any], prefix: str) -> StreamingResponse:
    output = generate_excel(data)
    filename = f"{prefix}_{_period_filename(data)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "Cache-Control": "no-cache",
        },
    )


# ---------------------------------------------------------------------------
# Routes — Current Week
# ---------------------------------------------------------------------------

@router.get(
    "/current-week/report",
    response_model=Dict[str, Any],
    dependencies=[Depends(require_analysis_roles)],
)
def get_current_week_report():
    """Full report for the current calendar week (Mon–Sun)."""
    return _report_for_current_week()


@router.get(
    "/current-week/report/excel",
    dependencies=[Depends(require_analysis_roles)],
)
def get_current_week_report_excel():
    """Excel export for the current calendar week."""
    return _excel_response(_report_for_current_week(), "report")


# ---------------------------------------------------------------------------
# Routes — 2025
# ---------------------------------------------------------------------------

@router.get(
    "/2025/report",
    response_model=Dict[str, Any],
    dependencies=[Depends(require_analysis_roles)],
)
def get_2025_report():
    """Full report for 2025."""
    return _report_for_year(2025)


@router.get(
    "/2025/report/excel",
    dependencies=[Depends(require_analysis_roles)],
)
def get_2025_report_excel():
    """Excel export for 2025."""
    return _excel_response(_report_for_year(2025), "report")


# ---------------------------------------------------------------------------
# Routes — 2026 yearly
# ---------------------------------------------------------------------------

@router.get(
    "/2026/report",
    response_model=Dict[str, Any],
    dependencies=[Depends(require_analysis_roles)],
)
def get_2026_report():
    """Full report for 2026."""
    return _report_for_year(2026)


@router.get(
    "/2026/report/excel",
    dependencies=[Depends(require_analysis_roles)],
)
def get_2026_report_excel():
    """Excel export for 2026."""
    return _excel_response(_report_for_year(2026), "report")


# ---------------------------------------------------------------------------
# Routes — 2026 monthly  (/2026/{month}/report, month = 1–12)
# NOTE: registered after the static /2026/report routes so FastAPI
#       matches those first.
# ---------------------------------------------------------------------------

def _validate_month(month: int):
    if month < 1 or month > 12:
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail="Month must be between 1 and 12",
        )


@router.get(
    "/2026/{month}/report",
    response_model=Dict[str, Any],
    dependencies=[Depends(require_analysis_roles)],
)
def get_2026_monthly_report(month: int):
    """Full report for a specific month of 2026."""
    _validate_month(month)
    return _report_for_month(2026, month)


@router.get(
    "/2026/{month}/report/excel",
    dependencies=[Depends(require_analysis_roles)],
)
def get_2026_monthly_report_excel(month: int):
    """Excel export for a specific month of 2026."""
    _validate_month(month)
    return _excel_response(_report_for_month(2026, month), "report")
